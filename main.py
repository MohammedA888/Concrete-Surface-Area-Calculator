# Must first install openpyxl library
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill

workbook = Workbook()
sheet = workbook.active

# This file name can be used as an example to see how the program works, accompanied by the 'Dataset Floor 3-5.txt' file
file_name = 'Dataset Floor 3-5'

sheet["A1"] = "Description"
sheet["B1"] = "Dim 1"
sheet["C1"] = "Dim 2"
sheet["D1"] = "Area"

for letter in ["A", "B", "C", "D"]:
    sheet[letter + "1"].fill = PatternFill(start_color='A6A6A6',
                                           end_color='A6A6A6',
                                           fill_type='solid')
    sheet[letter + "1"].alignment = Alignment(horizontal="center", vertical="center")

sum_so_far = 0
row = 1
with open(file_name + '.txt', 'r') as f:
    for line in f:
        row += 1
        if line[0] == '\n':
            row -= 1
            continue
        elif line[-2] == ':' and line[0].isnumeric() and line.find('Floor') != -1:
            if row != 2:
                row += 1
            sheet["A" + str(row)] = line[:-2]
            sheet.merge_cells("A" + str(row) + ":D" + str(row))
            sheet["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
            sheet["A" + str(row)].fill = PatternFill(start_color='00B0F0',
                                                     end_color='00B0F0',
                                                     fill_type='solid')
        elif line[0].isnumeric():
            sheet["A" + str(row)] = line[:-2]
            sheet.merge_cells("A" + str(row) + ":D" + str(row))
            sheet["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
            sheet["A" + str(row)].fill = PatternFill(start_color='FFFF00',
                                                     end_color='FFFF00',
                                                     fill_type='solid')
        else:
            col_loc = line.find(':')
            sheet["A" + str(row)] = line[:col_loc]
            if line[:col_loc].find('*') != -1:
                for letter in ["A", "B", "C", "D"]:
                    sheet[letter + str(row)].fill = PatternFill(start_color='FF0000',
                                                                end_color='FF0000',
                                                                fill_type='solid')

            sheet["A" + str(row)].alignment = Alignment(horizontal="center", vertical="center")

            x_loc = line.find('x', col_loc)
            
            sheet["B" + str(row)] = line[col_loc + 2:x_loc]
            sheet["B" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
            
            sheet["C" + str(row)] = line[x_loc + 1:-1]
            sheet["C" + str(row)].alignment = Alignment(horizontal="center", vertical="center")

            sheet["D" + str(row)] = str(int(float(line[col_loc + 2:x_loc]) * float(line[x_loc + 1:-1])))
            sheet["D" + str(row)].alignment = Alignment(horizontal="center", vertical="center")
            sum_so_far += int(float(line[col_loc + 2:x_loc]) * float(line[x_loc + 1:-1]))


sheet["A" + str(row + 2)] = "Total Concrete Surface Area (in^2)"
sheet.merge_cells("A" + str(row + 2)+ ":C" + str(row + 2))
sheet["D" + str(row + 2)] = sum_so_far
for letter in ["A", "D"]:
    sheet[letter + str(row + 2)].alignment = Alignment(horizontal="center", vertical="center")
    sheet[letter + str(row + 2)].fill = PatternFill(start_color='E26B0A',
                                                    end_color='E26B0A',
                                                    fill_type='solid')

sheet.column_dimensions["A"].width = 54
workbook.save(filename = file_name + ".xlsx")
